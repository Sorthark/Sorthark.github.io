"""Generate product showcase HTML from 茶叶咖啡商品信息.xlsx and extract product images."""
from __future__ import annotations

import json
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from html import escape
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL_GLOB = "*.xlsx"
LOGO_NAME = "品牌logo.jpg"
ASSETS_PRODUCTS = ROOT / "assets" / "products"
DISPIMG_RE = re.compile(r'DISPIMG\("([^"]+)"', re.I)

NS = {
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

BADGE_PATTERNS = [
    (re.compile(r"【推荐】|【数量拍"), "推荐", "product-badge"),
    (re.compile(r"【尝鲜|试饮|试鲜|品鉴"), "尝鲜", "product-badge-new"),
    (re.compile(r"【限购|拍1发"), "热销", "product-badge"),
    (re.compile(r"礼盒|送礼|伴手礼"), "礼盒", "product-badge-featured"),
]

CATEGORY_RULES = [
    (r"龙井|碧螺春|毛峰|绿茶|安吉白|峨眉|雨前|明前|冷泡.*绿", "绿茶"),
    (r"茉莉|花茶|窨", "花茶"),
    (r"普洱|小青柑|沱茶|熟茶|生茶|饼茶", "普洱"),
    (r"大红袍|岩茶|乌龙|铁观音|肉桂|单丛|鸭屎香", "乌龙"),
    (r"红茶|滇红|正山小种|金萱|桐木", "红茶"),
    (r"白茶|寿眉|白牡丹", "白茶"),
    (r"挂耳|咖啡豆|罗布斯塔|曼特宁|胶囊|冻干|黑咖啡|即饮|拿铁", "精品咖啡"),
    (r"白咖啡", "白咖啡"),
    (r"速溶|三合一|二合一", "速溶咖啡"),
    (r"组合|茗茶|礼盒|糕点|鹧鸪|陈皮", "组合/其他"),
]

UNIT_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(g|kg|ml|条|包|罐|瓶|粒|饼|盒|袋|装|克|毫升)(?:\s*/\s*(\w+))?",
    re.I,
)

SHEET_SLUG = {"茶叶": "tea", "咖啡": "coffee"}


def find_excel() -> Path:
    for path in sorted(ROOT.glob(EXCEL_GLOB)):
        if path.name.startswith("~$"):
            continue
        return path
    raise FileNotFoundError("No .xlsx file found in project root")


def load_dispimg_map(xlsx: Path) -> dict[str, str]:
    """Map WPS DISPIMG image IDs to zip paths like xl/media/image39.png."""
    with zipfile.ZipFile(xlsx) as zf:
        rels: dict[str, str] = {}
        rel_root = ET.fromstring(zf.read("xl/_rels/cellimages.xml.rels"))
        for rel in rel_root:
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid and "media/" in target:
                rels[rid] = "xl/" + target.replace("../", "")

        id_map: dict[str, str] = {}
        ci_root = ET.fromstring(zf.read("xl/cellimages.xml"))
        for cell_image in ci_root.findall(".//etc:cellImage", NS):
            name_el = cell_image.find(".//xdr:cNvPr", NS)
            blip = cell_image.find(".//a:blip", NS)
            if name_el is None or blip is None:
                continue
            img_id = name_el.attrib.get("name", "")
            embed = blip.attrib.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
            )
            if img_id and embed in rels:
                id_map[img_id] = rels[embed]
    return id_map


def parse_dispimg_id(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    match = DISPIMG_RE.search(text)
    return match.group(1) if match else None


def load_products(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        items = []
        for row in rows[1:]:
            if not row or row[1] is None:
                continue
            record = dict(zip(header, row))
            name = str(record.get("商品名称", "")).strip()
            if not name:
                continue
            price = record.get("价格")
            try:
                price_val = float(price) if price is not None else None
            except (TypeError, ValueError):
                price_val = None
            pid = record.get("商品序号")
            try:
                pid_int = int(pid) if pid is not None else len(items) + 1
            except (TypeError, ValueError):
                pid_int = len(items) + 1
            image_raw = record.get("商品图片") or record.get("图片")
            items.append(
                {
                    "id": pid_int,
                    "name": name,
                    "price": price_val,
                    "image_id": parse_dispimg_id(image_raw),
                }
            )
        sheets[sheet_name] = items
    wb.close()
    return sheets


def extract_product_images(
    xlsx: Path, sheets: dict[str, list[dict]], id_map: dict[str, str]
) -> dict[tuple[str, int], str]:
    """Extract images to assets/products/{tea|coffee}/{id}.ext; return relative web paths."""
    if ASSETS_PRODUCTS.exists():
        shutil.rmtree(ASSETS_PRODUCTS)
    ASSETS_PRODUCTS.mkdir(parents=True, exist_ok=True)

    path_by_key: dict[tuple[str, int], str] = {}
    needed: dict[tuple[str, int], str] = {}
    for sheet, items in sheets.items():
        slug = SHEET_SLUG.get(sheet, sheet)
        for item in items:
            img_id = item.get("image_id")
            if not img_id or img_id not in id_map:
                continue
            needed[(slug, item["id"])] = id_map[img_id]

    with zipfile.ZipFile(xlsx) as zf:
        for (slug, pid), zip_path in needed.items():
            ext = Path(zip_path).suffix.lower() or ".jpeg"
            out_dir = ASSETS_PRODUCTS / slug
            out_dir.mkdir(parents=True, exist_ok=True)
            out_file = out_dir / f"{pid:03d}{ext}"
            with zf.open(zip_path) as src, open(out_file, "wb") as dst:
                shutil.copyfileobj(src, dst)
            rel = out_file.relative_to(ROOT).as_posix()
            path_by_key[(slug, pid)] = rel
            item_ref = next(
                (it for it in sheets.get("茶叶" if slug == "tea" else "咖啡", []) if it["id"] == pid),
                None,
            )
            if item_ref is None:
                other = "咖啡" if slug == "tea" else "茶叶"
                item_ref = next(
                    (it for it in sheets.get(other, []) if it["id"] == pid),
                    None,
                )
    return path_by_key


def infer_subtype(name: str) -> str:
    for pattern, label in CATEGORY_RULES:
        if re.search(pattern, name):
            return label
    return "其他"


def extract_unit(name: str) -> str:
    matches = UNIT_RE.findall(name)
    if not matches:
        return ""
    num, unit, extra = matches[-1]
    unit = unit.lower().replace("克", "g").replace("毫升", "ml")
    suffix = f"/{extra}" if extra else ""
    return f"/ {num}{unit}{suffix}"


def detect_badge(name: str, price: float | None, sheet: str, index: int) -> tuple[str, str] | None:
    for pattern, label, css in BADGE_PATTERNS:
        if pattern.search(name):
            return label, css
    if price is not None and price >= 500:
        return "高端", "product-badge-featured"
    if index == 0:
        return "热销", "product-badge"
    return None


def short_desc(name: str) -> str:
    clean = re.sub(r"^渡南洋\s*", "", name).strip()
    clean = re.sub(r"【[^】]+】", "", clean).strip()
    if len(clean) <= 48:
        return clean
    return clean[:45] + "…"


def product_image_html(sheet: str, item: dict, image_paths: dict[tuple[str, int], str]) -> str:
    slug = SHEET_SLUG.get(sheet, sheet)
    key = (slug, item["id"])
    rel = image_paths.get(key)
    alt = escape(item["name"])
    if rel:
        return (
            f'<img src="{escape(rel)}" alt="{alt}" loading="lazy" '
            f'decoding="async" width="400" height="300">'
        )
    fallback = "product-image-tea-1" if sheet == "茶叶" else "product-image-coffee-1"
    return f'<div class="product-image-fallback {fallback}" aria-hidden="true"></div>'


def render_card(
    item: dict, sheet: str, index: int, image_paths: dict[tuple[str, int], str]
) -> str:
    name = item["name"]
    price = item["price"]
    subtype = infer_subtype(name)
    prefix = "东方茶" if sheet == "茶叶" else "南洋咖"
    category_class = (
        "product-category product-category--coffee"
        if sheet == "咖啡"
        else "product-category"
    )
    category = f"{prefix} · {subtype}"
    title = escape(name)
    desc = escape(short_desc(name))
    unit = extract_unit(name)
    unit_html = f" <small>{escape(unit)}</small>" if unit else ""
    price_html = f"¥{price:g}{unit_html}" if price is not None else "价格面议"
    badge = detect_badge(name, price, sheet, index)
    badge_html = ""
    if badge:
        label, css = badge
        badge_html = f'\n              <span class="product-badge {css}">{escape(label)}</span>'
    featured = (
        " product-card-featured"
        if badge and badge[1] == "product-badge-featured" and price and price >= 200
        else ""
    )
    img_inner = product_image_html(sheet, item, image_paths)
    return f"""          <article class="product-card{featured}">
            <div class="product-image">{badge_html}
              {img_inner}
            </div>
            <div class="product-body">
              <span class="{category_class}">{escape(category)}</span>
              <h3 class="product-name">{title}</h3>
              <p class="product-desc">{desc}</p>
              <div class="product-footer">
                <span class="product-price">{price_html}</span>
                <a href="#contact" class="product-link">咨询购买</a>
              </div>
            </div>
          </article>"""


def render_section(
    sheet: str, items: list[dict], image_paths: dict[tuple[str, int], str]
) -> str:
    label = "茶叶系列" if sheet == "茶叶" else "咖啡系列"
    eyebrow = "东方茶" if sheet == "茶叶" else "南洋咖"
    cards = "\n".join(
        render_card(item, sheet, i, image_paths) for i, item in enumerate(items)
    )
    return f"""        <div class="product-group" data-category="{escape(sheet)}">
          <h3 class="product-group-title">{eyebrow} · {label}</h3>
          <p class="product-group-count">共 {len(items)} 款商品</p>
          <div class="product-grid">
{cards}
          </div>
        </div>"""


def build_products_html(
    sheets: dict[str, list[dict]], image_paths: dict[tuple[str, int], str]
) -> str:
    sections = []
    if "茶叶" in sheets:
        sections.append(render_section("茶叶", sheets["茶叶"], image_paths))
    if "咖啡" in sheets:
        sections.append(render_section("咖啡", sheets["咖啡"], image_paths))
    sections_html = "\n".join(sections)
    return f"""    <section id="products" class="products section">
      <div class="container">
        <header class="section-header">
          <span class="section-label">商品展示</span>
          <h2 class="section-title">渡南洋精选茶咖</h2>
          <p class="section-desc">一杯茶香一缕咖醇🌴，解锁琼南洋跨境共生新机遇</p>
        </header>

{sections_html}
      </div>
    </section>"""


def patch_index(html: str, products_section: str) -> str:
    start = html.index('    <section id="products"')
    end = html.index('    <section id="contact"')
    return html[:start] + products_section + "\n\n" + html[end:]


def patch_brand_logo(html: str) -> str:
    logo = LOGO_NAME
    if not (ROOT / logo).is_file():
        return html

    nav_logo = f"""      <a href="#" class="logo">
        <img src="{logo}" alt="茶承东方 · 咖汇南洋" class="logo-img" width="44" height="44">
        <span class="logo-text">茶承东方 · 咖汇南洋</span>
      </a>"""

    html = re.sub(
        r'      <a href="#" class="logo">.*?</a>',
        nav_logo,
        html,
        count=1,
        flags=re.DOTALL,
    )

    hero_insert = f"""        <img src="{logo}" alt="" class="hero-brand-logo" width="120" height="120">
"""
    if "hero-brand-logo" not in html:
        html = html.replace(
            '        <p class="hero-eyebrow">',
            hero_insert + '        <p class="hero-eyebrow">',
            1,
        )

    footer_logo = f"""      <div class="footer-brand">
        <img src="{logo}" alt="茶承东方 · 咖汇南洋" class="logo-img footer-logo-img" width="56" height="56">
        <p>茶承东方，咖汇南洋</p>
      </div>"""
    html = re.sub(
        r'      <div class="footer-brand">.*?</div>',
        footer_logo,
        html,
        count=1,
        flags=re.DOTALL,
    )

    if f'href="{logo}"' not in html and 'rel="icon"' not in html:
        html = html.replace(
            '  <link rel="stylesheet" href="styles.css">',
            f'  <link rel="icon" href="{logo}" type="image/jpeg">\n'
            '  <link rel="stylesheet" href="styles.css">',
            1,
        )
    return html


def main() -> None:
    excel = find_excel()
    id_map = load_dispimg_map(excel)
    sheets = load_products(excel)
    image_paths = extract_product_images(excel, sheets, id_map)
    products_html = build_products_html(sheets, image_paths)

    index_path = ROOT / "index.html"
    html = index_path.read_text(encoding="utf-8")
    html = patch_index(html, products_html)
    html = patch_brand_logo(html)
    index_path.write_text(html, encoding="utf-8")

    meta = {
        "excel": excel.name,
        "tea": len(sheets.get("茶叶", [])),
        "coffee": len(sheets.get("咖啡", [])),
        "images_extracted": len(image_paths),
        "logo": LOGO_NAME if (ROOT / LOGO_NAME).is_file() else None,
    }
    print(json.dumps(meta, ensure_ascii=False))


if __name__ == "__main__":
    main()
